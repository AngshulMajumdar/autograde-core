from __future__ import annotations

import json
from pathlib import Path
from typing import List

import openpyxl

from autograde.extractors.base import BaseExtractor
from autograde.models import Artifact, EvidenceObject, Location


def _infer_type(values):
    vals = [v for v in values if v not in (None, "")]
    if not vals:
        return "empty"
    numeric = 0
    sample = vals[:50]
    for v in sample:
        try:
            float(v)
            numeric += 1
        except Exception:
            pass
    return "numeric" if numeric >= max(1, len(sample) * 0.7) else "categorical"


class DatasetExtractor(BaseExtractor):
    extractor_id = "dataset_extractor_v1"
    supported_types = {"dataset"}

    def extract(self, artifact: Artifact) -> List[EvidenceObject]:
        path = Path(artifact.storage_path)
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return self._extract_xlsx(artifact, path)
        if suffix == ".json":
            return self._extract_json(artifact, path)
        return self._extract_tsv(artifact, path)

    def _extract_tsv(self, artifact: Artifact, path: Path) -> List[EvidenceObject]:
        rows = [line.split("\t") for line in path.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()]
        header = rows[0] if rows else []
        body = rows[1:]
        schema = {h: _infer_type([r[i] if i < len(r) else None for r in body]) for i, h in enumerate(header)}
        return [self._dataset_evidence(artifact, path.name, "tsv_dataset", len(body), len(header), schema, header)]

    def _extract_json(self, artifact: Artifact, path: Path) -> List[EvidenceObject]:
        obj = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
        rows = obj if isinstance(obj, list) else obj.get("data", []) if isinstance(obj, dict) else []
        header = sorted({k for row in rows[:200] if isinstance(row, dict) for k in row.keys()})
        schema = {h: _infer_type([row.get(h) for row in rows if isinstance(row, dict)]) for h in header}
        return [self._dataset_evidence(artifact, path.name, "json_dataset", len(rows), len(header), schema, header)]

    def _extract_xlsx(self, artifact: Artifact, path: Path) -> List[EvidenceObject]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(x) if x is not None else "" for x in (rows[0] if rows else [])]
        body = rows[1:]
        schema = {h: _infer_type([r[i] if i < len(r) else None for r in body]) for i, h in enumerate(header) if h}
        wb.close()
        return [
            self._dataset_evidence(
                artifact,
                path.name,
                "xlsx_dataset",
                len(body),
                len([h for h in header if h]),
                schema,
                header,
            )
        ]

    def _dataset_evidence(self, artifact: Artifact, file_name: str, subtype: str, row_count: int, column_count: int, schema, header):
        return EvidenceObject(
            evidence_id=f"{artifact.artifact_id}_dataset_1",
            submission_id=artifact.submission_id,
            artifact_id=artifact.artifact_id,
            modality="dataset",
            subtype=subtype,
            content=None,
            structured_content={
                "row_count": row_count,
                "column_count": column_count,
                "header": header[:50],
                "schema": schema,
                "target_candidates": [h for h in header if str(h).lower() in {"label", "target", "y", "class"}],
            },
            preview=file_name,
            location=Location(file=file_name),
            confidence=0.96,
            extractor_id=self.extractor_id,
            tags=["dataset", "schema"],
        )
